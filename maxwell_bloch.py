import numpy as np
import matplotlib.pyplot as plt
import pickle
from cmath import sqrt, exp, pi
from scipy.integrate import complex_ode
from random import uniform
from openpyxl import Workbook
from tqdm import tqdm

num_a = 50
num_sigma_d = num_a
# initial conditions
a0 = 0.01
sigma0 = pow(10, -10)
d0 = 0.0

time_limit = 5 * 10 ** 3
t_standby_fourier = time_limit / 2

omega_0 = 3 * pow(10, -3)
Nat = 10 ** 5
gamma_sigma = 0.1
omega_sigma = 1
gamma_d = 10 ** -6
gamma_p = 100 * gamma_d


class MBSystem:
    def __init__(self):
        self.delta = np.ndarray(shape=(num_a,), dtype=np.complex128)
        self.gamma = np.ndarray(shape=(num_a,), dtype=np.complex128)

        self.omega = np.ndarray(shape=(num_a, num_sigma_d), dtype=np.complex128)

        self.sigma = np.ndarray(shape=(num_sigma_d,), dtype=np.complex128)
        self.d = np.ndarray(shape=(num_sigma_d,), dtype=np.complex128)
        self.a = np.ndarray(shape=(num_a,), dtype=np.complex128)

        self.time = []
        self.quantities = {}

        self.ode_init = []

        self.fourier_time = []
        self.fourier_quantities = {}

    def save_consts(self):
            # unfinished
        # wb = Workbook()
        # ws1 = wb.active
        # ws1.title = "omega"

        # for row in range(self.omega.shape[1]):
        #     for col in range(self.omega.shape[0]):
        #         _ = ws1.cell(column=col + 1, row=row + 1, value=str(self.omega[col][row]))

        # ws2 = wb.create_sheet(title="gamma")

        # for i in range(self.gamma.shape[0]):
        #     _ = ws2.cell(column=col + 1, row=row + 1, value=str(self.gamma[i]))

        # ws3 = wb.create_sheet(title="gamma")

        # for i in range(self.delta.shape[0]):
        #     _ = ws3.cell(column=col + 1, row=row + 1, value=str(self.delta[i]))

        # wb.save(filename='consts.xlsx')

        pickle.dump((self.gamma, self.delta, self.omega), open("consts_MB.p", "wb"))

    def generate_consts(self):
        for i in range(num_a):
            alpha = uniform(2, 4)
            for j in range(num_sigma_d):
                self.omega[i][j] = omega_0 * exp(-abs(i - j) * alpha)
                self.omega[j][i] = omega_0 * exp(-abs(i - j) * alpha)

        for i in range(num_a):
            self.delta[i] = uniform(- gamma_sigma, + gamma_sigma)
            self.gamma[i] = gamma_sigma * uniform(0.5, 1)

        self.save_consts()

    def load_consts(self):
        self.gamma, self.delta, self.omega = pickle.load(open("consts.p", "rb"))

    def initial_conditions(self):
        for i in range(num_a):
            self.ode_init.append(a0)
        for i in range(num_sigma_d):
            self.ode_init.append(sigma0)
        for i in range(num_sigma_d):
            self.ode_init.append(d0)

        for i in range(num_a):
            self.a[i] = a0
        for i in range(num_sigma_d):
            self.sigma[i] = sigma0
            self.d[i] = d0

    def ode_sys(self):
        def func(t, y):
            temp_sigma = np.ndarray(shape=(num_sigma_d,), dtype=np.complex128)
            temp_d = np.ndarray(shape=(num_sigma_d,), dtype=np.complex128)
            temp_a = np.ndarray(shape=(num_a,), dtype=np.complex128)
            temp_arr = np.ndarray(shape=(num_a + 2 * num_sigma_d,), dtype=np.complex128)

            for ii in range(num_a):
                temp_a[ii] = y[ii]
            for ii in range(num_sigma_d):
                temp_sigma[ii] = y[ii + num_a]
            for ii in range(num_sigma_d):
                temp_d[ii] = y[ii + num_a + num_sigma_d]

            temp1 = Nat * np.dot(self.omega, temp_sigma)
            temp2 = np.dot(temp_a, np.conj(self.omega))
            temp3 = np.dot(np.conj(temp_a), self.omega)
            temp4 = np.dot(temp_a, np.conj(self.omega))

            for ii in range(num_a):
                temp_arr[ii] = (-1 * 1j * self.delta[ii] - self.gamma[ii]) * temp_a[ii] - 1j * temp1[ii]
            for ii in range(num_sigma_d):
                temp_arr[ii + num_a] = -1 * gamma_sigma * temp_sigma[ii] + 1j * temp2[ii] * temp_d[ii]
            for ii in range(num_sigma_d):
                temp_arr[ii + num_a + num_sigma_d] = (gamma_p - gamma_d) - (gamma_p + gamma_d) * temp_d[
                    ii] + 2 * 1j * (temp3[ii] * temp_sigma[ii] - temp4[ii] * np.conj(temp_sigma[ii]))
            return temp_arr

        return func

    def append_quantities(self, t):
        def append_quantity(tag, quantity):
            try:
                self.quantities[tag].append(quantity)
            except KeyError:
                self.quantities[tag] = [quantity]

        def append_quantity_for_fourier(tag, quantity):
            try:
                self.fourier_quantities[tag].append(quantity)
            except KeyError:
                self.fourier_quantities[tag] = [quantity]

        self.time.append(t)

        append_quantity('field_avg', np.average(self.a).real)
        append_quantity('field_first', self.a[0].real)
        if t > t_standby_fourier:
            self.fourier_time.append(t)
            append_quantity_for_fourier('field_avg', np.average(self.a))
            append_quantity_for_fourier('field_first', self.a[0])

    def plot_dynamic(self):
        plt.figure(figsize=(18, 8))
        font = {'size': 16}
        plt.rc('font', **font)
        plt.xlabel('t')
        for tag in self.quantities:
            plt.plot(self.time, self.quantities[tag], label=tag)
        plt.grid(True)
        plt.legend()
        plt.show()

    def plot_fourier(self):
        fourier_transformed = {}
        n = len(self.fourier_time)
        n = int(n * 100)
        for tag in self.fourier_quantities:
            values = self.fourier_quantities[tag]
            fourier_transformed[tag] = abs(np.fft.ifft(values, n=n))
        freq = np.fft.fftfreq(n)
        freq *= 2 * pi

        plt.figure(figsize=(18, 8))
        font = {'size': 16}
        plt.rc('font', **font)
        plt.xlabel('ω')
        for tag in self.fourier_quantities:
            plt.plot(freq, fourier_transformed[tag], label=tag)
        plt.grid(True)
        plt.legend()
        plt.show()

    def decompose_ode_sol(self, sol):
        for i in range(num_a):
            self.a[i] = sol[i]
        for i in range(num_sigma_d):
            self.sigma[i] = sol[i + num_a]
            self.d[i] = sol[i + num_a + num_sigma_d]

    def solve_equations(self):
        time_stamps = np.concatenate((np.linspace(0, time_limit / 100, num=10 ** 3),
                                      np.linspace(time_limit / 100, time_limit, num=10 ** 3)))

        self.initial_conditions()
        ode = complex_ode(self.ode_sys()).set_integrator('dopri5', method='bdf')
        ode.set_initial_value(self.ode_init)
        for t in tqdm(time_stamps[1:]):
            dt = t - ode.t
            solution = ode.integrate(ode.t + dt)
            self.decompose_ode_sol(sol=solution)
            self.append_quantities(t)
